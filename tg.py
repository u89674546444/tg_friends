import os
import logging
import json
from datetime import datetime
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
    ConversationHandler,
)
from fpdf import FPDF
import openpyxl

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Состояния для ConversationHandler
REQUESTING_PHONE, REQUESTING_FULL_NAME, SELECTING_HOUSE, SELECTING_WORK_TYPE, RECEIVING_PHOTO_BEFORE, RECEIVING_PHOTO_AFTER, CONFIRMING_COMPLETION, CONTINUING_WORK, SELECTING_UNFINISHED_JOB = range(9)
# Путь к шрифту
FONT_PATH = "/Users/nikolajusakov/PycharmProjects/PythonProject/dejavu-fonts-ttf-2.37/ttf/DejaVuSans.ttf"

# Путь к Excel-файлам
USERS_EXCEL_PATH = "/Users/nikolajusakov/PycharmProjects/PythonProject/users.xlsx"
HOUSES_EXCEL_PATH = "/Users/nikolajusakov/PycharmProjects/PythonProject/Houses.xlsx"
WORKS_EXCEL_PATH = "/Users/nikolajusakov/PycharmProjects/PythonProject/works.xlsx"

# Путь к файлам с данными пользователей и незавершенными работами
USERS_DATA_PATH = "/Users/nikolajusakov/PycharmProjects/PythonProject/users_data.json"
UNFINISHED_JOBS_PATH = "/Users/nikolajusakov/PycharmProjects/PythonProject/unfinished_jobs.json"


# Проверка наличия шрифта
if not os.path.exists(FONT_PATH):
    logger.error(f"Шрифт {FONT_PATH} не найден.")
    exit(1)

# Проверка и создание необходимых файлов
def initialize_files():
    if not os.path.exists(USERS_EXCEL_PATH):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Users"
        sheet.append(["Phone", "Full Name"])
        workbook.save(USERS_EXCEL_PATH)

    if not os.path.exists(USERS_DATA_PATH):
        with open(USERS_DATA_PATH, "w", encoding="utf-8") as file:
            json.dump({}, file)

    if not os.path.exists(UNFINISHED_JOBS_PATH):
        with open(UNFINISHED_JOBS_PATH, "w", encoding="utf-8") as file:
            json.dump({}, file)

initialize_files()

# Функции для работы с данными
def save_user_to_excel(phone, full_name):
    workbook = openpyxl.load_workbook(USERS_EXCEL_PATH)
    sheet = workbook.active
    sheet.append([phone, full_name])
    workbook.save(USERS_EXCEL_PATH)

def get_user_data(user_id):
    try:
        with open(USERS_DATA_PATH, "r", encoding="utf-8") as file:
            users_data = json.load(file)
        return users_data.get(str(user_id))
    except Exception as e:
        logger.error(f"Ошибка при получении данных пользователя: {e}")
        return None

def save_user_data(user_id, phone, full_name):
    try:
        with open(USERS_DATA_PATH, "r", encoding="utf-8") as file:
            users_data = json.load(file)

        users_data[str(user_id)] = {"phone": phone, "full_name": full_name}

        with open(USERS_DATA_PATH, "w", encoding="utf-8") as file:
            json.dump(users_data, file, ensure_ascii=False, indent=4)
    except Exception as e:
        logger.error(f"Ошибка при сохранении данных пользователя: {e}")

def add_unfinished_job(user_id, house_number, full_name, house_full_name, work_type, photo_before):
    try:
        with open(UNFINISHED_JOBS_PATH, "r", encoding="utf-8") as file:
            unfinished_jobs = json.load(file)

        if str(user_id) not in unfinished_jobs:
            unfinished_jobs[str(user_id)] = []

        unfinished_jobs[str(user_id)].append({
            "house_number": house_number,
            "full_name": full_name,
            "house_full_name": house_full_name,
            "work_type": work_type,
            "photo_before": photo_before,
            "status": "Не завершено"
        })

        with open(UNFINISHED_JOBS_PATH, "w", encoding="utf-8") as file:
            json.dump(unfinished_jobs, file, ensure_ascii=False, indent=4)
    except Exception as e:
        logger.error(f"Ошибка при добавлении незавершенной работы: {e}")

def get_unfinished_jobs(user_id):
    try:
        with open(UNFINISHED_JOBS_PATH, "r", encoding="utf-8") as file:
            unfinished_jobs = json.load(file)
        return unfinished_jobs.get(str(user_id), [])
    except Exception as e:
        logger.error(f"Ошибка при получении списка незавершенных работ: {e}")
        return []

def remove_unfinished_job(user_id, house_number):
    try:
        with open(UNFINISHED_JOBS_PATH, "r", encoding="utf-8") as file:
            unfinished_jobs = json.load(file)

        if str(user_id) in unfinished_jobs:
            unfinished_jobs[str(user_id)] = [job for job in unfinished_jobs[str(user_id)] if
                                             job["house_number"] != house_number]

        with open(UNFINISHED_JOBS_PATH, "w", encoding="utf-8") as file:
            json.dump(unfinished_jobs, file, ensure_ascii=False, indent=4)
    except Exception as e:
        logger.error(f"Ошибка при удалении незавершенной работы: {e}")

# Обработчик команды /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.info("Пользователь начал взаимодействие с ботом.")
    user_id = update.message.from_user.id
    user_data = get_user_data(user_id)

    if user_data:
        # Пользователь уже авторизован
        unfinished_jobs = get_unfinished_jobs(user_id)

        if unfinished_jobs:
            # Есть незавершенные работы
            jobs_list = "\n".join(
                [f"{i + 1}. {job['work_type']} ({job['house_full_name']})" for i, job in enumerate(unfinished_jobs)]
            )
            await update.message.reply_text(f"Ваши незавершенные работы:\n{jobs_list}\nВведите номер работы для продолжения:")
            return SELECTING_UNFINISHED_JOB
        else:
            # Нет незавершенных работ
            await update.message.reply_text("У вас нет незавершенных работ. Выберите номер дома:")
            return SELECTING_HOUSE
    else:
        # Пользователь не авторизован
        await update.message.reply_text("Пожалуйста, введите ваш номер телефона:")
        return REQUESTING_PHONE

# Обработчик ввода номера телефона
async def request_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    phone = update.message.text
    context.user_data["phone"] = phone  # Сохраняем phone в context.user_data
    await update.message.reply_text("Введите ваше ФИО:")
    return REQUESTING_FULL_NAME

# Обработчик ввода ФИО
async def request_full_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    full_name = update.message.text
    context.user_data["full_name"] = full_name  # Сохраняем full_name в context.user_data
    save_user_to_excel(context.user_data["phone"], full_name)
    save_user_data(update.message.from_user.id, context.user_data["phone"], full_name)
    await update.message.reply_text("Авторизация завершена. Выберите номер дома:")
    return SELECTING_HOUSE

# Обработчик полного имени дома
def get_house_full_name(house_number):
    try:
        workbook = openpyxl.load_workbook(HOUSES_EXCEL_PATH)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            if str(row[0]) == str(house_number):  # Сравниваем как строки
                return row[1]  # Возвращаем полное название
        return None  # Если дом не найден
    except Exception as e:
        logger.error(f"Ошибка при чтении файла домов: {e}")
        return None
# Обработчик выбора дома
async def select_house(update: Update, context: ContextTypes.DEFAULT_TYPE):
    selected_house = update.message.text
    context.user_data["selected_house"] = selected_house
    house_full_name = get_house_full_name(selected_house)
    if not house_full_name:
        await update.message.reply_text("Дом не найден. Пожалуйста, выберите номер дома из списка.")
        return SELECTING_HOUSE

    # Получаем список типов работ
    work_types = get_work_types()
    if not work_types:
        await update.message.reply_text("Ошибка: список типов работ пуст.")
        return ConversationHandler.END

    # Создаем клавиатуру с кнопками для выбора типа работ
    keyboard = [[KeyboardButton(work_type)] for work_type in work_types]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True)

    await update.message.reply_text(
        f"Выбран дом: {house_full_name}. Выберите тип работы:",
        reply_markup=reply_markup
    )
    return SELECTING_WORK_TYPE

    # Получаем список типов работ
    work_types = get_work_types()
    if not work_types:
        await update.message.reply_text("Ошибка: список типов работ пуст.")
        return ConversationHandler.END

    # Создаем клавиатуру с кнопками для выбора типа работ
    keyboard = [[KeyboardButton(work_type)] for work_type in work_types]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True)

    await update.message.reply_text(
        f"Выбран дом: {house_full_name}. Выберите тип работы:",
        reply_markup=reply_markup
    )
    return SELECTING_WORK_TYPE



# Путь к файлу Excel
WORKS_EXCEL_PATH = 'works.xlsx'


# Функция для чтения данных из Excel и записи во второй колонке в текстовый файл
def write_descriptions_to_text_file():
    try:
        # Загружаем рабочую книгу
        workbook = openpyxl.load_workbook(WORKS_EXCEL_PATH)
        sheet = workbook.active

        # Открываем текстовый файл для записи
        with open('descriptions.txt', 'w', encoding='utf-8') as text_file:
            # Проходим по всем строкам, начиная со второй (пропуская заголовок)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Записываем значение из второй колонки в текстовый файл
                text_file.write(f"{row[1]}\n")  # Предполагаем, что описание во втором столбце

    except Exception as e:
        print(f"Ошибка при обработке файла: {e}")


# Вызов функции
write_descriptions_to_text_file()
# обработчик типа работ в файле
def get_work_types():
    try:
        workbook = openpyxl.load_workbook(WORKS_EXCEL_PATH)
        sheet = workbook.active
        work_types = []
        descriptions = []

        # Читаем данные из файла
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
            work_types.append(row[0])  # Первый столбец — тип работы
            descriptions.append(row[1])  # Второй столбец — описание

        # Записываем второй столбец в текстовый файл
        with open('descriptions.txt', 'w', encoding='utf-8') as file:
            for description in descriptions:
                file.write(f"{description}\n")

        return work_types
    except Exception as e:
        logger.error(f"Ошибка при чтении файла работ: {e}")
        return []
# Обработчик выбора типа работ
async def select_work_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    work_type = update.message.text
    context.user_data["work_type"] = work_type
    await update.message.reply_text("Пришлите фото до начала работ.")
    return RECEIVING_PHOTO_BEFORE

# Обработчик получения фото "до начала работ"
async def handle_photo_before(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Проверяем, есть ли необходимые данные в context.user_data
    if "full_name" not in context.user_data or "selected_house" not in context.user_data or "work_type" not in context.user_data:
        await update.message.reply_text("Ошибка: данные пользователя не найдены. Пожалуйста, начните с команды /start.")
        return ConversationHandler.END

    # Остальная логика функции
    photo = update.message.photo[-1]
    file = await photo.get_file()
    photo_before = f"photo_before_{context.user_data['selected_house']}.jpg"
    await file.download_to_drive(photo_before)
    context.user_data["photo_before"] = photo_before
    logger.info(f"Фото 'до начала работ' сохранено: {photo_before}.")

    # Добавляем незавершенную работу
    add_unfinished_job(
        update.message.from_user.id,
        context.user_data["selected_house"],
        context.user_data["full_name"],
        get_house_full_name(context.user_data["selected_house"]),
        context.user_data["work_type"],
        photo_before
    )

    # Предлагаем кнопку для добавления второго фото или завершения работы
    keyboard = [[KeyboardButton("Добавить 2 фото")], [KeyboardButton("Завершить работу")]]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True)
    await update.message.reply_text("Выберите действие:", reply_markup=reply_markup)
    return RECEIVING_PHOTO_AFTER

# Обработчик получения фото "после окончания работ"
async def handle_photo_after(update: Update, context: ContextTypes.DEFAULT_TYPE):
    photo = update.message.photo[-1]
    file = await photo.get_file()
    photo_after = f"photo_after_{context.user_data['selected_house']}.jpg"
    await file.download_to_drive(photo_after)
    context.user_data["photo_after"] = photo_after
    logger.info(f"Фото 'после окончания работ' сохранено: {photo_after}.")

    # Формируем PDF
    def create_pdf(user_data):
        try:
            # Проверяем наличие необходимых данных
            if "house_full_name" not in user_data or "work_type" not in user_data or "photo_before" not in user_data or "photo_after" not in user_data:
                logger.error("Недостаточно данных для создания PDF.")
                return None

            pdf = FPDF()
            pdf.add_page()
            pdf.add_font('DejaVu', '', FONT_PATH, uni=True)
            pdf.set_font('DejaVu', '', 12)

            # Добавляем данные в PDF
            pdf.cell(200, 10, txt="Отчет о выполненной работе", ln=True, align='C')
            pdf.cell(200, 10, txt=f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
            pdf.cell(200, 10, txt=f"Полный адрес: {user_data['house_full_name']}", ln=True)
            pdf.cell(200, 10, txt=f"Тип работ: {user_data['work_type']}", ln=True)
            pdf.cell(200, 10, txt="Фото до начала работ:", ln=True)
            pdf.image(user_data["photo_before"], x=10, y=pdf.get_y(), w=100)
            pdf.cell(200, 10, txt="Фото после окончания работ:", ln=True)
            pdf.image(user_data["photo_after"], x=10, y=pdf.get_y(), w=100)

            # Сохраняем PDF
            pdf_path = f"report_{user_data['selected_house']}.pdf"
            pdf.output(pdf_path)
            return pdf_path
        except Exception as e:
            logger.error(f"Ошибка при создании PDF: {e}")
            return None

# Обработчик выбора незавершенной работы
async def select_unfinished_job(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    unfinished_jobs = get_unfinished_jobs(user_id)

    try:
        job_index = int(update.message.text) - 1
        if 0 <= job_index < len(unfinished_jobs):
            selected_job = unfinished_jobs[job_index]

            # Сохраняем данные выбранной работы в context.user_data
            context.user_data["selected_house"] = selected_job["house_number"]
            context.user_data["work_type"] = selected_job["work_type"]
            context.user_data["house_full_name"] = selected_job["house_full_name"]
            context.user_data["photo_before"] = selected_job["photo_before"]

            # Отправляем первое фото
            await update.message.reply_photo(selected_job["photo_before"], caption=f"Продолжаем работу: {context.user_data['work_type']} ({selected_job['house_full_name']})")
            await update.message.reply_text("Пришлите фото после окончания работ.")
            return RECEIVING_PHOTO_BEFORE
        else:
            await update.message.reply_text("Неверный номер работы. Попробуйте снова.")
            return CONTINUING_WORK
    except ValueError:
        await update.message.reply_text("Неверный ввод. Введите номер работы (число).")
        return CONTINUING_WORK

# Функция для создания PDF
def create_pdf(user_data):
    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.add_font('DejaVu', '', FONT_PATH, uni=True)
        pdf.set_font('DejaVu', '', 12)

        # Добавляем данные в PDF
        pdf.cell(200, 10, txt="Отчет о выполненной работе", ln=True, align='C')
        pdf.cell(200, 10, txt=f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
        pdf.cell(200, 10, txt=f"Полный адрес: {user_data['house_full_name']}", ln=True)
        pdf.cell(200, 10, txt=f"Тип работ: {user_data['work_type']}", ln=True)
        pdf.cell(200, 10, txt="Фото до начала работ:", ln=True)
        pdf.image(user_data["photo_before"], x=10, y=pdf.get_y(), w=100)
        pdf.cell(200, 10, txt="Фото после окончания работ:", ln=True)
        pdf.image(user_data["photo_after"], x=10, y=pdf.get_y(), w=100)

        # Сохраняем PDF
        pdf_path = f"report_{user_data['selected_house']}.pdf"
        pdf.output(pdf_path)
        return pdf_path
    except Exception as e:
        logger.error(f"Ошибка при создании PDF: {e}")
        return None

# Обработчик отмены
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.info("Пользователь отменил действие.")
    await update.message.reply_text("Действие отменено.")
    return ConversationHandler.END

# Регистрация команд
async def post_init(application: Application):
    await application.bot.set_my_commands([
        ("start", "Начать работу с ботом"),
        ("unfinished_jobs", "Показать незавершенные работы"),
    ])
# Обработчик подтверждения завершения работы
async def confirm_completion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_choice = update.message.text
    user_id = update.message.from_user.id

    if user_choice == "Завершить работу":
        # Удаляем незавершенную работу
        remove_unfinished_job(user_id, context.user_data["selected_house"])

        # Обновляем статус в папке
        house_full_name = get_house_full_name(context.user_data["selected_house"])
        save_files_to_folder(
            context.user_data["selected_house"],
            context.user_data["full_name"],
            house_full_name,
            context.user_data["work_type"],
            context.user_data["photo_before"],
            is_completed=True
        )

        await update.message.reply_text("Работа завершена. Спасибо!")
    else:
        await update.message.reply_text("Работа сохранена как незавершенная. Вы можете продолжить позже.")

    return ConversationHandler.END

# Обработчик продолжения работы
async def continue_work(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    unfinished_jobs = get_unfinished_jobs(user_id)

    if update.message.text == "Новая работа":
        await update.message.reply_text("Выберите номер дома:")
        return SELECTING_HOUSE
    elif update.message.text == "Продолжить работу":
        if unfinished_jobs:
            # Формируем пронумерованный список работ
            jobs_list = "\n".join(
                [f"{i + 1}. {job['work_type']} ({job['house_full_name']})" for i, job in enumerate(unfinished_jobs)]
            )
            await update.message.reply_text(
                f"Ваши незавершенные работы:\n{jobs_list}\nВведите номер работы для продолжения:")
            return SELECTING_UNFINISHED_JOB
        else:
            await update.message.reply_text("У вас нет незавершенных работ. Выберите номер дома:")
            return SELECTING_HOUSE

# Основная функция
def main():
    # Создаем экземпляр Application
    application = Application.builder().token('8127498518:AAFzskJYwY0-gkF2FdjJbtY1YcjZVd88wGs').post_init(post_init).build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            REQUESTING_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, request_phone)],
            REQUESTING_FULL_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, request_full_name)],
            SELECTING_HOUSE: [MessageHandler(filters.TEXT & ~filters.COMMAND, select_house)],
            SELECTING_WORK_TYPE: [MessageHandler(filters.TEXT & ~filters.COMMAND, select_work_type)],
            RECEIVING_PHOTO_BEFORE: [MessageHandler(filters.PHOTO, handle_photo_before)],
            RECEIVING_PHOTO_AFTER: [MessageHandler(filters.PHOTO, handle_photo_after)],
            CONFIRMING_COMPLETION: [MessageHandler(filters.TEXT & ~filters.COMMAND, confirm_completion)],
            CONTINUING_WORK: [MessageHandler(filters.TEXT & ~filters.COMMAND, continue_work)],
            SELECTING_UNFINISHED_JOB: [MessageHandler(filters.TEXT & ~filters.COMMAND, select_unfinished_job)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    application.add_handler(conv_handler)
    application.run_polling()

if __name__ == '__main__':
    main()